from openpyxl import Workbook, load_workbook
import os, random

def create_or_load_workbook(path):
    if not os.path.exists(path):
        wb=Workbook()
        wb.save(path)
    else:
        print("Path alread exists!")

    return load_workbook(path)

def get_data_from_excel(sheet):
    data={'sheet_count':sheet.max_row,'questions':[],'options':[],'answer':""}
    for i in range(1,data['sheet_count']+1):
        data['questions'].append(sheet[f'A{i}'].value)
        data['options'].append(sheet[f'B{i}'].value)
        data['answer']= sheet[f'C{i}'].value
    return data


def display_question(data):
    print('ddddddddddddddddddddddddddd',data)
    print(data['sheet_count'])
    random_question=random.randint(0,data['sheet_count'])-1
    print(random_question)
    index_opt=["1.","2.","3.","4."]
    opt=(data['options'][random_question]).split(',')

    print('Here is your question, try to select the option correctly.\n')

    print(data['questions'][random_question])

    for i in range(min(len(opt), len(index_opt))):
        print(f"{index_opt[i]} {opt[i]}")
    print('-----------',type(opt),"---",opt)
    answer_index=(opt.index(data['answer']))-1
    original_answer=[answer_index, data['answer']]
    
    return original_answer

    # for i in range(min(len(opt), len(index_opt))):
#     print(f"{index_opt[i]}{opt[i]}")
    


def add_quesiont_in_excel():
    print("i) Here you can add number of questions.")
    print("ii) Mention the number of options you going to add(max option is 4).")
    print("iii) Then add your options one by one.")
    print("iv) Finally mention your answers\n")

    question=input("Write the question here->")
    number_of_opt=int(input("Enter the options->"))
    options=""
    for i in range(number_of_opt):
        get_opt=input(f'Enter option number {i+1}:')
        options += f'{get_opt.strip()}'
    answer=input('Now add the correct answer from the option->')
    if not answer in options:
        return 'check your answer, Try again!'
    return question, options, answer

def save_questions(wb_load,sheet, path ,add):
    sheet_count=(sheet.max_row)
    sheet[f'A{sheet_count+1}']= add[0]
    sheet[f'B{sheet_count+1}']= add[1]
    sheet[f'C{sheet_count+1}']= add[2]
    wb_load.save(path)

    return "Saved the question successfully!"

def get_user_choice():
    return int(input("Enter the option:"))

def check_answer(selected_option,original_answer):
    if selected_option == original_answer[0]:
        print('You have chosen the correct answer:', original_answer[1])
    else:
        print(f'You have chosen the wrong answer. The correct answer is {original_answer[1]}')

    
    
    

    


def main():
    path=r"C:\Users\Kishore\OneDrive\Desktop\Front_end\pyproject\weather_app\Weather_app\quiz_game\quiz.xlsx"
    wb_load=create_or_load_workbook(path)
    sheet=wb_load.active
    
    # add=add_quesiont_in_excel()
    # print(add)
    # s1=save_questions(wb_load, sheet,path, add)
    # print(s1)
    # get_choice=get_user_choice()
    # opt = display_question(sheet)
    # answer = opt.index(sheet['C1'].value) + 1
    a=get_data_from_excel(sheet)
    print(a)

    b=display_question(a)
    selected_option=get_user_choice()
    c=check_answer(selected_option,b)
    print(c)




main()




# from openpyxl import Workbook, load_workbook
# import os



# path=r"C:\Users\Kishore\OneDrive\Desktop\Front_end\pyproject\weather_app\Weather_app\quiz_game\quiz.xlsx"
# if not os.path.exists(path):
#     wb=Workbook()
#     wb.save(path)
# else:
#     print('path exist already')


# wb_=load_workbook(path)
# sheet=wb_.active


# sheet['A1'] = "What is your name?"
# sheet['B1'] = "Male,Female"
# sheet['C1'] = "Male"

# wb_.save(path)

# index_opt=["1. ","2. ","3. ","4. "]
# opt=(sheet['B1'].value).split(',')

# # print(sheet['A1'].value)
# # Assuming opt is a list or iterable
# for i in range(min(len(opt), len(index_opt))):
#     print(f"{index_opt[i]}{opt[i]}")
# selected_option=int(input('Enter the option: '))
# answer=opt.index(sheet['c1'].value)+1

# if selected_option==answer:
#     print('You have choosen a right answer',sheet['c1'].value)
# else:
#     print('You have choosen a wrong answer')




# import random

# # Sample quiz questions
# questions = {
#     'Python': [
#         {
#             'question': 'What is the capital of Brazil?',
#             'options': ['Brasília', 'Rio de Janeiro', 'São Paulo', 'Salvador'],
#             'answer': 'Brasília'
#         },
#         # Add more questions for other topics/categories
#     ]
# }

# def display_question(question_data):
#     print(question_data['question'])
#     for i, option in enumerate(question_data['options'], start=1):
#         print(f"{i}. {option}")

# def validate_answer(question_data, user_answer):
#     return user_answer == question_data['answer']

# def quiz():
#     topic = input("Choose a topic for the quiz (e.g., Python): ")
#     topic_questions = questions.get(topic, [])
#     if not topic_questions:
#         print("Invalid topic. Please choose a valid topic.")
#         return

#     score = 0
#     for question_data in random.sample(topic_questions, k=len(topic_questions)):
#         display_question(question_data)
#         user_answer = input("Enter your answer (1-4): ")
#         if user_answer.isdigit() and 1 <= int(user_answer) <= len(question_data['options']):
#             if validate_answer(question_data, question_data['options'][int(user_answer) - 1]):
#                 print("Correct!")
#                 score += 1
#             else:
#                 print("Incorrect.")
#         else:
#             print("Invalid input. Please enter a number between 1 and 4.")

#     print(f"\nQuiz completed! Your score: {score}/{len(topic_questions)}")

# if __name__ == "__main__":
#     quiz()
